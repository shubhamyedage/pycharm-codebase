from os import listdir, getcwd
from os.path import join, isfile, exists
import pandas as pd
from openpyxl import load_workbook, Workbook
from os.path import join, expanduser


class FileUtils(object):
    def __init__(self):
        self.root_path = join(expanduser('~'), 'report-analysis', 'source')

    def get_files(self, src_dir_path):
        path = join(self.root_path, src_dir_path)
        file_list = [f for f in listdir(path) if isfile(join(path, f))]
        return file_list

    def read_csv(self, src_dir_path, f):
        path = join(self.root_path, src_dir_path, f)
        data = pd.read_csv(path, index_col=0, header=1)
        data.dropna(how='all')
        data.fillna(0, inplace=True)
        data.replace('(%)', '', inplace=True, regex=True)
        return data

    def write_csv(self, out_dir_path, data):
        path = join(self.root_path, out_dir_path)
        data.to_csv(path, index=False)

    def read_excel(self, src_dir_path, f, sheets_list):
        path = join(self.root_path, src_dir_path, f)
        data = pd.read_excel(path, sheetname=sheets_list, header=0, index_col=0)
        return data

    def write_excel(self, out_dir_path, df, sheet_name):
        path = join(self.root_path, out_dir_path)
        book = Workbook()
        if not exists(path):
            book.save(path)
        book = load_workbook(path)
        # book.remove_sheet(book.worksheets[0])
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book
        for i in book.worksheets:
            if i.title == sheet_name:
                book.remove_sheet(i)
        df.to_excel(writer, sheet_name=sheet_name)
        writer.save()
